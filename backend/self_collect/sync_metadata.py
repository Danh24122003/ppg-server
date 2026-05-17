"""
Đọc CSV trong collected_data/ → append metadata vào subjects_metadata.xlsx.

S01ÍNH SÁS01:
  - File CSV đã được link (csv_file column trong xlsx khớp tên) → SKIP.
  - File CSV chưa link → APPEND new row ở cuối sheet 'Subjects'.
  - KHÔNG bao giờ ghi đè cell có sẵn data (manual input giữ nguyên).
  - Cell formula (bmi, sbp_baseline_mean, dbp_baseline_mean) ghi formula đúng row.
  - Manual columns (sbp_baseline_1/2, dbp_baseline_1/2, medication_name,
    signal_quality_notes) để TRỐNG — bạn fill tay sau khi xem Omron.

USAGE:
  python sync_metadata.py            # apply (with auto-backup .bak)
  python sync_metadata.py --dry-run  # preview only, không ghi gì

BACKUP: tự tạo subjects_metadata.xlsx.bak trước khi save.

YÊU CẦU: Excel ĐÓNG file xlsx (script detect lock file ~$ và exit nếu đang mở).
"""
from __future__ import annotations

import argparse
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import openpyxl

HERE = Path(__file__).resolve().parent
XLSX_PATH = HERE / "subjects_metadata.xlsx"
LOCK_PATH = HERE / "~$subjects_metadata.xlsx"
COLLECTED_DIR = HERE.parent.parent / "collected_data"
SHEET_NAME = "Subjects"


# ============================================================
# Parse CSV comment header → dict
# ============================================================
@dataclass
class CSVMeta:
    csv_file: str
    subject_id: str
    age: Optional[int] = None
    sex: Optional[str] = None
    height_cm: Optional[float] = None
    weight_kg: Optional[float] = None
    sbp_baseline: Optional[int] = None  # mean từ console
    dbp_baseline: Optional[int] = None  # mean từ console
    sbp_post: Optional[int] = None
    dbp_post: Optional[int] = None
    hypertension: Optional[str] = None
    medication: Optional[str] = None
    smoking: Optional[str] = None
    finger: Optional[str] = None
    cuff_arm: Optional[str] = None
    session_date: Optional[str] = None
    fs: Optional[int] = None
    duration_s: Optional[float] = None
    notes: Optional[str] = None


def _parse_kv_line(line: str) -> Dict[str, str]:
    """Parse '# k1=v1,k2=v2,...' → dict. Bỏ '# ' prefix, split ','."""
    body = line.lstrip("#").strip()
    out: Dict[str, str] = {}
    for chunk in body.split(","):
        if "=" in chunk:
            k, v = chunk.split("=", 1)
            out[k.strip()] = v.strip()
    return out


def _to_int(v: Optional[str]) -> Optional[int]:
    if v is None or v == "" or v.upper() == "NA":
        return None
    try:
        return int(float(v))
    except ValueError:
        return None


def _to_float(v: Optional[str]) -> Optional[float]:
    if v is None or v == "" or v.upper() == "NA":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def parse_csv_header(path: Path) -> Optional[CSVMeta]:
    """Đọc 5 dòng comment đầu CSV → CSVMeta. None nếu format sai."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = []
            for _ in range(8):  # đủ buffer cho 5 dòng `#` + 1 header CSV
                line = f.readline()
                if not line:
                    break
                if line.startswith("#"):
                    lines.append(line)
                else:
                    break
    except OSError as exc:
        print(f"  [WARN] không đọc được {path.name}: {exc}")
        return None

    if not lines:
        print(f"  [WARN] {path.name}: không có comment header → skip")
        return None

    kv: Dict[str, str] = {}
    for line in lines:
        kv.update(_parse_kv_line(line))

    sid = kv.get("subject_id", "").strip()
    if not sid:
        print(f"  [WARN] {path.name}: thiếu subject_id → skip")
        return None

    return CSVMeta(
        csv_file=path.name,
        subject_id=sid,
        age=_to_int(kv.get("age")),
        sex=kv.get("sex"),
        height_cm=_to_float(kv.get("height_cm")),
        weight_kg=_to_float(kv.get("weight_kg")),
        sbp_baseline=_to_int(kv.get("sbp_baseline")),
        dbp_baseline=_to_int(kv.get("dbp_baseline")),
        sbp_post=_to_int(kv.get("sbp_post")),
        dbp_post=_to_int(kv.get("dbp_post")),
        hypertension=kv.get("hypertension"),
        medication=kv.get("medication"),
        smoking=kv.get("smoking"),
        finger=kv.get("finger"),
        cuff_arm=kv.get("cuff_arm"),
        session_date=kv.get("session_date"),
        fs=_to_int(kv.get("fs")),
        duration_s=_to_float(kv.get("duration_s")),
        notes=kv.get("notes"),
    )


# ============================================================
# Sync logic
# ============================================================
def get_existing_csv_files(ws) -> set[str]:
    """Trả set các csv_file đã có trong xlsx (đã được sync trước đó)."""
    headers = [c.value for c in ws[1]]
    if "csv_file" not in headers:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' thiếu cột 'csv_file'")
    col_idx = headers.index("csv_file") + 1
    out: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=col_idx).value
        if v is not None and str(v).strip():
            out.add(str(v).strip())
    return out


def find_first_empty_row(ws) -> int:
    """Trả index row đầu tiên 'thực sự trống' (tất cả cell None hoặc chỉ có
    subject_id placeholder S###).

    Rule: row được coi là 'có thể fill' nếu mọi cột TRỪ subject_id đều None.
    Nếu không có row nào như vậy → return ws.max_row + 1 (append cuối).
    """
    headers = [c.value for c in ws[1]]
    sid_col = headers.index("subject_id") + 1
    for r in range(2, ws.max_row + 1):
        non_sid_filled = False
        for c in range(1, len(headers) + 1):
            if c == sid_col:
                continue
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                non_sid_filled = True
                break
        if not non_sid_filled:
            return r
    return ws.max_row + 1


def write_row(ws, row_idx: int, meta: CSVMeta, headers: List[str]) -> None:
    """Ghi metadata CSV vào row_idx. Manual cols để None (cell trống).
    Formula cols ghi formula đúng row index."""
    # column letter helpers
    def col_letter(name: str) -> str:
        idx = headers.index(name) + 1
        return openpyxl.utils.get_column_letter(idx)

    def set_cell(name: str, value) -> None:
        if name not in headers:
            return
        col = headers.index(name) + 1
        # Chỉ ghi nếu cell hiện tại trống (an toàn cho rewrite)
        cur = ws.cell(row=row_idx, column=col).value
        if cur is None or str(cur).strip() == "":
            ws.cell(row=row_idx, column=col).value = value

    # Auto-fill từ CSV
    set_cell("subject_id", meta.subject_id)
    set_cell("session_date", meta.session_date.split("T")[0] if meta.session_date else None)
    set_cell("age", meta.age)
    set_cell("sex", meta.sex)
    set_cell("height_cm", meta.height_cm)
    set_cell("weight_kg", meta.weight_kg)
    set_cell("hypertension", meta.hypertension)
    set_cell("medication", meta.medication)
    set_cell("smoking", meta.smoking)
    set_cell("finger", meta.finger)
    set_cell("cuff_arm", meta.cuff_arm)
    set_cell("sbp_post", meta.sbp_post)
    set_cell("dbp_post", meta.dbp_post)
    set_cell("csv_file", meta.csv_file)
    set_cell("duration_s", meta.duration_s)
    set_cell("fs_hz", meta.fs)
    set_cell("general_notes", meta.notes if meta.notes else None)

    # Formula cells (chỉ set nếu trống — không ghi đè)
    h_col = col_letter("height_cm")
    w_col = col_letter("weight_kg")
    bmi_formula = f"=IFERROR(ROUND({w_col}{row_idx}/(({h_col}{row_idx}/100)^2),1),\"\")"
    set_cell("bmi", bmi_formula)

    if "sbp_baseline_1" in headers and "sbp_baseline_2" in headers:
        n_col = col_letter("sbp_baseline_1")
        p_col = col_letter("sbp_baseline_2")
        set_cell("sbp_baseline_mean", f"=IFERROR(ROUND(({n_col}{row_idx}+{p_col}{row_idx})/2,0),\"\")")
    if "dbp_baseline_1" in headers and "dbp_baseline_2" in headers:
        o_col = col_letter("dbp_baseline_1")
        q_col = col_letter("dbp_baseline_2")
        set_cell("dbp_baseline_mean", f"=IFERROR(ROUND(({o_col}{row_idx}+{q_col}{row_idx})/2,0),\"\")")

    # Manual cols để trống cho user fill: sbp_baseline_1/2, dbp_baseline_1/2,
    # medication_name, signal_quality_notes — KHÔNG set_cell.


# ============================================================
# Main
# ============================================================
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dry-run", action="store_true", help="preview only, không ghi xlsx")
    args = ap.parse_args()

    print("=" * 64)
    print("  Sync metadata: collected_data/ → subjects_metadata.xlsx")
    print("=" * 64)

    # Lock check
    if LOCK_PATH.exists():
        print(f"\n[ERROR] Excel đang mở file ({LOCK_PATH.name} tồn tại).")
        print("        Đóng Excel rồi chạy lại.")
        return 1

    if not XLSX_PATH.exists():
        print(f"[ERROR] Không tìm thấy {XLSX_PATH}")
        return 1

    if not COLLECTED_DIR.exists():
        print(f"[ERROR] Không tìm thấy {COLLECTED_DIR}")
        return 1

    # Read all CSVs
    csv_files = sorted(COLLECTED_DIR.glob("*.csv"))
    if not csv_files:
        print(f"\n[INFO] {COLLECTED_DIR} không có file CSV. Done.")
        return 0

    print(f"\nFound {len(csv_files)} CSV(s) trong collected_data/:")
    metas: List[CSVMeta] = []
    for p in csv_files:
        m = parse_csv_header(p)
        if m is None:
            continue
        metas.append(m)
        print(f"  - {p.name}  (subject_id={m.subject_id}, fs={m.fs}, duration={m.duration_s}s)")

    if not metas:
        print("\n[INFO] không có CSV hợp lệ. Done.")
        return 0

    # Open xlsx
    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] xlsx thiếu sheet '{SHEET_NAME}'. Sheets: {wb.sheetnames}")
        return 1
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]

    existing = get_existing_csv_files(ws)
    print(f"\nĐã có {len(existing)} CSV được link trong xlsx: {sorted(existing) if existing else '(none)'}")

    # Plan actions
    to_append: List[CSVMeta] = []
    skipped: List[str] = []
    for m in metas:
        if m.csv_file in existing:
            skipped.append(m.csv_file)
        else:
            to_append.append(m)

    print(f"\nKế hoạch:")
    print(f"  SKIP (đã sync): {len(skipped)}")
    for f in skipped:
        print(f"    - {f}")
    print(f"  APPEND (mới):   {len(to_append)}")
    for m in to_append:
        print(f"    - {m.csv_file} → row mới (subject_id={m.subject_id})")

    if not to_append:
        print("\n[OK] Tất cả CSV đã sync, không có gì để append.")
        return 0

    if args.dry_run:
        print("\n[DRY-RUN] không ghi xlsx. Bỏ --dry-run để apply thật.")
        return 0

    # Backup
    bak = XLSX_PATH.with_suffix(".xlsx.bak")
    shutil.copy2(XLSX_PATH, bak)
    print(f"\n[BACKUP] {bak.name}")

    # Append
    for m in to_append:
        # Ưu tiên fill row trống có sẵn (S### placeholder), fallback append cuối
        target_row = find_first_empty_row(ws)
        write_row(ws, target_row, m, headers)
        print(f"  [WRITE] row {target_row} ← {m.csv_file}")

    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        print(f"\n[ERROR] Không save được — Excel có thể vừa mở file.")
        print(f"        Backup vẫn còn ở {bak}")
        return 1

    print(f"\n[OK] Saved {XLSX_PATH.name}: {len(to_append)} row(s) appended, {len(skipped)} skipped.")
    print("     Mở Excel, fill các cột tay: sbp_baseline_1/2, dbp_baseline_1/2,")
    print("     medication_name, signal_quality_notes.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
